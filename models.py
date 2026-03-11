from __future__ import annotations

from dataclasses import dataclass
from difflib import SequenceMatcher
from enum import Enum
from pathlib import Path
from typing import Any
import json
import re
import unicodedata


class PoliticalParty(Enum):
    """Political parties represented in this project."""

    PLN = "Partido Liberacion Nacional (PLN)"
    PPSD = "Partido Progreso Social Democratico (PPSD)"
    PUSC = "Partido Unidad Social Cristiana (PUSC)"
    NUEVA_REPUBLICA = "Partido Nueva Republica"
    FRENTE_AMPLIO = "Frente Amplio"
    PLP = "Partido Liberal Progresista (PLP)"
    INDEPENDENT = "Independent"

    @classmethod
    def from_text(cls, raw: str) -> PoliticalParty:
        """Resolve a party from enum key, acronym or full display name."""
        normalized = cls._normalize_text(raw)
        alias_map = {
            "pln": cls.PLN,
            "ppsd": cls.PPSD,
            "pusc": cls.PUSC,
            "nueva_republica": cls.NUEVA_REPUBLICA,
            "nueva republica": cls.NUEVA_REPUBLICA,
            "frente_amplio": cls.FRENTE_AMPLIO,
            "frente amplio": cls.FRENTE_AMPLIO,
            "plp": cls.PLP,
            "independent": cls.INDEPENDENT,
            "independiente": cls.INDEPENDENT,
            "ind": cls.INDEPENDENT,
        }
        if normalized in alias_map:
            return alias_map[normalized]

        for member in cls:
            if normalized == cls._normalize_text(member.name) or normalized == cls._normalize_text(member.value):
                return member

        raise ValueError(f"unknown political party: {raw}")

    @staticmethod
    def _normalize_text(value: str) -> str:
        base = unicodedata.normalize("NFKD", value.strip().lower())
        return "".join(ch for ch in base if not unicodedata.combining(ch))


@dataclass(frozen=True)
class Voter:
    """A deputy (voter) with name and political party."""

    name: str
    party: PoliticalParty
    seat_number: int | None = None

    def to_dict(self) -> dict[str, Any]:
        return {
            "name": self.name,
            "party": self.party.value,
            "seat_number": self.seat_number,
        }


class VoteChoice(Enum):
    """Possible voting choices for one deputy."""

    IN_FAVOR = "in_favor"
    AGAINST = "against"
    ABSTENTION = "abstention"
    ABSENT = "absent"


class VotingSession:
    """Stores each deputy vote for one voting session."""

    def __init__(self, name: str, png_path: str | Path, deputies: list[Voter]) -> None:
        self.name = name.strip()
        self.png_path = str(png_path)
        self._votes: dict[Voter, VoteChoice] = {deputy: VoteChoice.ABSENT for deputy in deputies}

    @property
    def votes(self) -> dict[Voter, VoteChoice]:
        return dict(self._votes)

    def get_votes_by_seat(self) -> dict[int, VoteChoice]:
        votes_by_seat: dict[int, VoteChoice] = {}
        for deputy, choice in self._votes.items():
            if isinstance(deputy.seat_number, int) and deputy.seat_number > 0:
                votes_by_seat[deputy.seat_number] = choice
        return votes_by_seat

    def register_vote(self, deputy: Voter, choice: VoteChoice) -> None:
        if deputy not in self._votes:
            raise ValueError(f"deputy is not part of this voting session: {deputy.name}")
        self._votes[deputy] = choice

    def get_vote(self, deputy: Voter) -> VoteChoice:
        if deputy not in self._votes:
            raise ValueError(f"deputy is not part of this voting session: {deputy.name}")
        return self._votes[deputy]

    def process_results(self, extractor: Any | None = None) -> dict[str, list[str]]:
        """Process OCR output and map raw names to known deputies."""
        if extractor is None:
            from text_extractor import TextExtractor

            extractor = TextExtractor()

        raw_results = extractor.extract_results(self.png_path)
        if not isinstance(raw_results, dict):
            raise ValueError("extractor must return a dictionary")

        # Reset current session votes before processing a new OCR run.
        for deputy in self._votes:
            self._votes[deputy] = VoteChoice.ABSENT

        section_to_choice = {
            "a_favor": VoteChoice.IN_FAVOR,
            "en_contra": VoteChoice.AGAINST,
            "no_votacion": VoteChoice.ABSTENTION,
        }

        unmatched: dict[str, list[str]] = {key: [] for key in section_to_choice}
        assigned: set[Voter] = set()

        for section_key, choice in section_to_choice.items():
            raw_names = raw_results.get(section_key, [])
            if not isinstance(raw_names, list):
                continue

            for raw_name in raw_names:
                if not isinstance(raw_name, str):
                    continue

                deputy = self._match_deputy_name(raw_name, assigned)
                if deputy is None:
                    unmatched[section_key].append(raw_name)
                    continue

                self._votes[deputy] = choice
                assigned.add(deputy)

        return unmatched

    def _match_deputy_name(self, raw_name: str, assigned: set[Voter]) -> Voter | None:
        """Map one OCR name string to the closest deputy in this session."""
        query = self._normalize_person_name(raw_name)
        if not query:
            return None

        best_deputy: Voter | None = None
        best_score = 0.0
        for deputy in self._votes:
            if deputy in assigned:
                continue

            candidate = self._normalize_person_name(deputy.name)
            score = self._name_similarity_score(query, candidate)
            if score > best_score:
                best_score = score
                best_deputy = deputy

        if best_score < 0.58:
            return None
        return best_deputy

    @staticmethod
    def _name_similarity_score(a: str, b: str) -> float:
        """Blend character and token similarity for OCR-robust matching."""
        ratio_direct = SequenceMatcher(None, a, b).ratio()

        tokens_a = [tok for tok in a.split() if tok]
        tokens_b = [tok for tok in b.split() if tok]
        sorted_a = " ".join(sorted(tokens_a))
        sorted_b = " ".join(sorted(tokens_b))
        ratio_tokens = SequenceMatcher(None, sorted_a, sorted_b).ratio()

        set_a = set(tokens_a)
        set_b = set(tokens_b)
        overlap = len(set_a & set_b) / max(1, len(set_a | set_b))

        return (0.35 * ratio_direct) + (0.35 * ratio_tokens) + (0.30 * overlap)

    @staticmethod
    def _normalize_person_name(value: str) -> str:
        base = unicodedata.normalize("NFKD", value.lower())
        no_accents = "".join(ch for ch in base if not unicodedata.combining(ch))
        no_punct = re.sub(r"[^a-z0-9,\s]", " ", no_accents)
        collapsed = re.sub(r"\s{2,}", " ", no_punct).strip(" ,")
        if "," not in collapsed:
            return collapsed

        # OCR often returns "last_name, first_name", while assembly roster uses "first_name last_name".
        left, right = collapsed.split(",", 1)
        return f"{right.strip()} {left.strip()}".strip()

    def to_dict(self, group_by_parties: bool = False, group_by_vote: bool = False) -> dict[str, Any]:
        vote_counts = {
            "in_favor": 0,
            "against": 0,
            "abstention": 0,
            "absent": 0,
        }
        for choice in self._votes.values():
            vote_counts[choice.value] += 1

        payload: dict[str, Any] = {
            "name": self.name,
            "png_path": self.png_path,
            "total_deputies": len(self._votes),
            "vote_counts": vote_counts,
            "votes": [
                {
                    "deputy": deputy.to_dict(),
                    "choice": choice.value,
                }
                for deputy, choice in self._votes.items()
            ],
        }

        if group_by_parties:
            votes_by_party: dict[str, list[dict[str, Any]]] = {}
            for deputy, choice in self._votes.items():
                party_name = deputy.party.value
                votes_by_party.setdefault(party_name, []).append(
                    {
                        "deputy": deputy.to_dict(),
                        "choice": choice.value,
                    }
                )
            payload["votes_by_party"] = votes_by_party

        if group_by_vote:
            votes_by_choice: dict[str, list[dict[str, Any]]] = {
                "in_favor": [],
                "against": [],
                "abstention": [],
                "absent": [],
            }
            for deputy, choice in self._votes.items():
                votes_by_choice[choice.value].append(
                    {
                        "deputy": deputy.to_dict(),
                        "choice": choice.value,
                    }
                )
            payload["votes_by_choice"] = votes_by_choice

        return payload

    def to_excel(
        self,
        xlsx_path: str | Path,
        group_by_parties: bool = False,
        group_by_vote: bool = False,
    ) -> Path:
        try:
            from openpyxl import Workbook
        except ImportError as exc:
            raise ImportError(
                "openpyxl is required for Excel export. Install it with: "
                "python -m pip install openpyxl"
            ) from exc

        payload = self.to_dict(
            group_by_parties=group_by_parties,
            group_by_vote=group_by_vote,
        )

        workbook = Workbook()

        summary_sheet = workbook.active
        summary_sheet.title = "summary"
        summary_sheet.append(["field", "value"])
        summary_sheet.append(["name", payload["name"]])
        summary_sheet.append(["png_path", payload["png_path"]])
        summary_sheet.append(["total_deputies", payload["total_deputies"]])
        summary_sheet.append(["in_favor", payload["vote_counts"]["in_favor"]])
        summary_sheet.append(["against", payload["vote_counts"]["against"]])
        summary_sheet.append(["abstention", payload["vote_counts"]["abstention"]])
        summary_sheet.append(["absent", payload["vote_counts"]["absent"]])

        votes_sheet = workbook.create_sheet("votes")
        votes_sheet.append(["seat_number", "deputy_name", "party", "choice"])
        for vote in payload["votes"]:
            deputy = vote["deputy"]
            votes_sheet.append(
                [deputy["seat_number"], deputy["name"], deputy["party"], vote["choice"]]
            )

        if group_by_parties and "votes_by_party" in payload:
            by_party_sheet = workbook.create_sheet("by_party")
            by_party_sheet.append(["party", "seat_number", "deputy_name", "choice"])
            votes_by_party = payload["votes_by_party"]
            for party_name, party_votes in votes_by_party.items():
                for vote in party_votes:
                    deputy = vote["deputy"]
                    by_party_sheet.append(
                        [party_name, deputy["seat_number"], deputy["name"], vote["choice"]]
                    )

        if group_by_vote and "votes_by_choice" in payload:
            by_vote_sheet = workbook.create_sheet("by_vote")
            by_vote_sheet.append(["choice", "seat_number", "deputy_name", "party"])
            votes_by_choice = payload["votes_by_choice"]
            for choice_name, choice_votes in votes_by_choice.items():
                for vote in choice_votes:
                    deputy = vote["deputy"]
                    by_vote_sheet.append(
                        [choice_name, deputy["seat_number"], deputy["name"], deputy["party"]]
                    )

        output_path = Path(xlsx_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
        return output_path


# Backward-compatibility alias.
Votacion = VotingSession


class Assembly:
    """Assembly model containing only the list of deputies."""

    def __init__(self) -> None:
        self.deputies: list[Voter] = []

    def add_deputy(self, deputy: Voter) -> None:
        self.deputies.append(deputy)

    def get_seat_assignments(self) -> list[dict[str, Any]]:
        """Return assembly composition ordered by seat number."""
        assignments = sorted(self.deputies, key=lambda dep: (dep.seat_number or 10**9, dep.name))
        return [
            {
                "seat_number": deputy.seat_number,
                "name": deputy.name,
                "party": deputy.party.value,
            }
            for deputy in assignments
        ]

    def load_roster_from_json(self, json_path: str | Path) -> None:
        """Populate deputies from a JSON roster."""
        path = Path(json_path)
        if not path.exists():
            raise FileNotFoundError(f"roster file not found: {path}")

        data = json.loads(path.read_text(encoding="utf-8-sig"))
        deputies = data.get("deputies")
        if not isinstance(deputies, list):
            raise ValueError("JSON must contain a 'deputies' list")

        self.deputies.clear()
        for index, item in enumerate(deputies, start=1):
            if not isinstance(item, dict):
                raise ValueError(f"deputies[{index}] must be an object")

            name = str(item.get("name", "")).strip()
            if not name:
                raise ValueError(f"deputies[{index}] is missing 'name'")

            raw_party = item.get("party")
            if not isinstance(raw_party, str) or not raw_party.strip():
                raise ValueError(f"deputies[{index}] is missing 'party'")

            raw_seat = item.get("seat_number", index)
            if not isinstance(raw_seat, int) or raw_seat < 1:
                raise ValueError(f"deputies[{index}] has invalid 'seat_number'")

            party = PoliticalParty.from_text(raw_party)
            self.add_deputy(Voter(name=name, party=party, seat_number=raw_seat))

    def to_dict(self) -> dict[str, Any]:
        return {
            "total_deputies": len(self.deputies),
            "deputies": [deputy.to_dict() for deputy in self.deputies],
        }
